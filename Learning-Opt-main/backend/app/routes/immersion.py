import os
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO
import json
import traceback
from app import config  # DB execution helper

immersion_bp = Blueprint('immersion', __name__)

basedir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
TEMPLATE_PATH = os.path.join(basedir, "uploads", "templates", "grades2.xlsx")
UPLOAD_JSON_PATH = os.path.join(basedir, "backend", "app", "static", "excel", "uploaded_data.json")

print("Resolved TEMPLATE_PATH:", TEMPLATE_PATH)
print("Exists?", os.path.exists(TEMPLATE_PATH))


@immersion_bp.route("/fill-template", methods=["POST"])
def fill_template():
    if "file" not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400

    try:
       # Load uploaded Excel
        file.stream.seek(0)
        wb_uploaded = load_workbook(file.stream, data_only=True)
        ws_uploaded = wb_uploaded.active

        # ---------------------- Extract SCHOOL & BATCH ----------------------
        school = str(ws_uploaded["F1"].value or "").strip()
        batch = str(ws_uploaded["G1"].value or "").strip()


        # ---------------------- Read all student rows (start row 10) ----------------------
        data = []
        for row in ws_uploaded.iter_rows(min_row=10, values_only=True):
            if not any(row):
                continue
            data.append({
                "LAST_NAME": row[0] or "",
                "FIRST_NAME": row[1] or "",
                "MIDDLE_NAME": row[2] or "",
                "STRAND": row[3] or "",
                "DEPARTMENT": row[4] or "",
                "WI": float(row[5] or 0),
                "CO": float(row[6] or 0),
                "5S": float(row[7] or 0),
                "BO": float(row[8] or 0),
                "CBO": float(row[9] or 0),
                "SDG": float(row[10] or 0),
                "OHSA": float(row[11] or 0),
                "WE": float(row[12] or 0),
                "UJC": float(row[13] or 0),
                "ISO": float(row[14] or 0),
                "PO": float(row[15] or 0),
                "HR": float(row[16] or 0),
                "PERDEV": float(row[17] or 0),
                "SUPP": float(row[18] or 0),
                "DS": float(row[19] or 0)
            })

            # ✅ Attach school + batch to every entry
        for entry in data:
            entry["SCHOOL"] = school
            entry["BATCH"] = batch

        # ---------------------- Compute Totals & Grades ----------------------
        written_fields = ["WI", "CO", "5S", "BO", "CBO", "SDG"]
        performance_fields = ["OHSA", "WE", "UJC", "ISO", "PO", "HR", "PERDEV", "SUPP", "DS"]

        for entry in data:
            total_score = sum(entry[f] for f in written_fields + performance_fields)
            entry["TOTAL_SCORE"] = total_score

            entry["WRITTEN_RATING"] = round(sum(entry[f] for f in written_fields) / len(written_fields), 2)
            entry["PERFORMANCE_RATING"] = round(sum(entry[f] for f in performance_fields) / len(performance_fields), 2)

            # Assign FINAL_GRADE & REMARKS
            if total_score >= 90:
                entry["FINAL_GRADE"] = "A"
            elif total_score >= 80:
                entry["FINAL_GRADE"] = "B"
            elif total_score >= 70:
                entry["FINAL_GRADE"] = "C"
            elif total_score >= 60:
                entry["FINAL_GRADE"] = "D"
            else:
                entry["FINAL_GRADE"] = "F"

            entry["REMARKS"] = "Passed" if entry["FINAL_GRADE"] != "F" else "Failed"

        # ---------------------- Save to Database ----------------------
        # First insert into immersion_batches
        if school and batch:
            row = config.fetch_one(
                "SELECT id FROM immersion_batches WHERE school=%s AND batch=%s",
                (school, batch)
            )
            if not row:
                config.execute_query(
                    "INSERT INTO immersion_batches (school, batch) VALUES (%s, %s)",
                    (school, batch)
                )

        for entry in data:
            try:
                config.execute_query("""
                    INSERT INTO immersion_records (
                        last_name, first_name, middle_name, strand, department,
                        WI, CO, 5S, BO, CBO, SDG,
                        OHSA, WE, UJC, ISO, PO, HR,
                        PERDEV, SUPP, DS,
                        total_score, written_rating, performance_rating, final_grade, remarks
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s, %s
                    )
                """, (
                    entry["LAST_NAME"], entry["FIRST_NAME"], entry["MIDDLE_NAME"], entry["STRAND"], entry["DEPARTMENT"],
                    int(entry["WI"]), int(entry["CO"]), int(entry["5S"]), int(entry["BO"]), int(entry["CBO"]), int(entry["SDG"]),
                    int(entry["OHSA"]), int(entry["WE"]), int(entry["UJC"]), int(entry["ISO"]), int(entry["PO"]), int(entry["HR"]),
                    int(entry["PERDEV"]), int(entry["SUPP"]), int(entry["DS"]),
                    float(entry["TOTAL_SCORE"]), float(entry["WRITTEN_RATING"]), float(entry["PERFORMANCE_RATING"]),
                    entry["FINAL_GRADE"], entry["REMARKS"]
                ))
            except Exception as e:
                print(f"❌ DB insert failed for {entry['LAST_NAME']}, {entry['FIRST_NAME']}: {e}")

        # ---------------------- Save JSON for frontend ----------------------
        os.makedirs(os.path.dirname(UPLOAD_JSON_PATH), exist_ok=True)
        with open(UPLOAD_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        # ---------------------- Fill Excel Template ----------------------
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({"error": f"Template not found at {TEMPLATE_PATH}"}), 500

        wb_template = load_workbook(TEMPLATE_PATH)
        ws_template = wb_template.active
        start_row = 10

        # ✅ Insert school + batch into A8
        ws_template["A8"] = f"{school} - {batch}"    

        for idx, entry in enumerate(data):
            row = start_row + idx
            ws_template.cell(row=row, column=1, value=idx + 1)
            ws_template.cell(row=row, column=2, value=entry["LAST_NAME"])
            ws_template.cell(row=row, column=3, value=entry["FIRST_NAME"])
            ws_template.cell(row=row, column=4, value=entry["MIDDLE_NAME"])
            ws_template.cell(row=row, column=5, value=entry["STRAND"])
            ws_template.cell(row=row, column=6, value=entry["DEPARTMENT"])

        output = BytesIO()
        wb_template.save(output)
        output.seek(0)

        return jsonify({
            "message": "Data saved to DB and template filled successfully",
            "school": school,
            "batch": batch,
            "rows": data
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@immersion_bp.route("/data", methods=["GET"])
def get_immersion_data():
    if not os.path.exists(UPLOAD_JSON_PATH):
        return jsonify({"rows": []})
    with open(UPLOAD_JSON_PATH, "r", encoding="utf-8") as f:
        rows = json.load(f)
    return jsonify({"rows": rows})