import os
import json
import traceback
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from app import config

upload_bp = Blueprint("upload", __name__)

@upload_bp.route("/upload", methods=["POST"])
def upload_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400

    try:
       # Load workbook
        file.stream.seek(0)
        wb = load_workbook(file.stream, data_only=True)
        ws = wb.active

        # --- Extract SCHOOL & BATCH from F1 & G1 ---
        school = str(ws["F1"].value or "").strip()
        batch = str(ws["G1"].value or "").strip()

        # --- Extract students (start at row 10 after headers) ---
        students = []
        for row in ws.iter_rows(min_row=10, values_only=True):
            if not any(row):
                continue
            students.append({
                "last_name": row[0] or "",
                "first_name": row[1] or "",
                "middle_name": row[2] or "",
                "strand": row[3] or "",
                "department": row[4] or "",
                "wi": row[5] or 0,
                "co": row[6] or 0,
                "5s": row[7] or 0,
                "bo": row[8] or 0,
                "cbo": row[9] or 0,
                "sdg": row[10] or 0,
                "ohsa": row[11] or 0,
                "we": row[12] or 0,
                "ujc": row[13] or 0,
                "iso": row[14] or 0,
                "po": row[15] or 0,
                "hr": row[16] or 0,
                "perdev": row[17] or 0,
                "supp": row[18] or 0,
                "ds": row[19] or 0
            })

            

        # --- Compute totals & grades ---
        written_fields = ["wi", "co", "5s", "bo", "cbo", "sdg"]
        performance_fields = ["ohsa", "we", "ujc", "iso", "po", "hr", "perdev", "supp", "ds"]

        for stu in students:
            for key in written_fields + performance_fields:
                try:
                    stu[key] = float(stu[key] or 0)
                except ValueError:
                    stu[key] = 0.0

            total_score = sum(stu[k] for k in written_fields + performance_fields)
            stu["total_score"] = total_score

            stu["written_rating"] = round(sum(stu[k] for k in written_fields) / len(written_fields), 2)
            stu["performance_rating"] = round(sum(stu[k] for k in performance_fields) / len(performance_fields), 2)

            if total_score >= 90:
                stu["final_grade"] = "A"
            elif total_score >= 80:
                stu["final_grade"] = "B"
            elif total_score >= 70:
                stu["final_grade"] = "C"
            elif total_score >= 60:
                stu["final_grade"] = "D"
            else:
                stu["final_grade"] = "F"

            stu["remarks"] = "Passed" if stu["final_grade"] != "F" else "Failed"

        

        # --- Insert each student ---
        for stu in students:
            if not any(str(v).strip() for v in stu.values()):
                continue
            try:
                config.execute_query("""
                    INSERT INTO immersion_records (
                        last_name, first_name, middle_name, strand, department,
                        WI, CO, 5S, BO, CBO, SDG,
                        OHSA, WE, UJC, ISO, PO, HR,
                        PERDEV, SUPP, DS,
                        total_score, written_rating, performance_rating, final_grade, remarks
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s, %s
                    )
                """, (
                    stu["last_name"], stu["first_name"], stu["middle_name"], stu["strand"], stu["department"],
                    int(stu["wi"]), int(stu["co"]), int(stu["5s"]), int(stu["bo"]), int(stu["cbo"]), int(stu["sdg"]),
                    int(stu["ohsa"]), int(stu["we"]), int(stu["ujc"]), int(stu["iso"]), int(stu["po"]), int(stu["hr"]),
                    int(stu["perdev"]), int(stu["supp"]), int(stu["ds"]),
                    float(stu["total_score"]), float(stu["written_rating"]), float(stu["performance_rating"]),
                    stu["final_grade"], stu["remarks"]
                ))
            except Exception as e:
                print(f"❌ DB insert failed for {stu['last_name']}, {stu['first_name']}: {e}")

        return jsonify({
            "message": "Upload processed successfully",
            "school": school,
            "batch": batch,
            "count": len(students)
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
